from flask import Blueprint, render_template, request, abort, redirect, url_for, current_app
import json
import re
from .nlp.extractor import SKILLS_CONHECIDAS, extrair_skills_dos_projetos
from openpyxl import load_workbook
import os
from openai import OpenAI

main = Blueprint('main', __name__)

COMPATIBILIDADE_LIMIAR = 30

def get_db():
    db = current_app.config.get('DB')
    if db is not None:
        return db
    uri = os.getenv('MONGODB_URI')
    if not uri:
        current_app.config['DB_ERROR'] = 'MONGODB_URI ausente no .env.'
        return None
    try:
        from pymongo import MongoClient
        kwargs = {'serverSelectionTimeoutMS': 5000}
        try:
            import certifi
            kwargs['tlsCAFile'] = certifi.where()
            kwargs['tls'] = True
        except Exception:
            pass
        client = MongoClient(uri, **kwargs)
        client.admin.command('ping')
        db_name = os.getenv('MONGODB_DB', 'talentflow')
        db = client[db_name]
        current_app.config['DB'] = db
        current_app.config['DB_ERROR'] = None
        return db
    except Exception as e:
        current_app.config['DB_ERROR'] = str(e)
        return None

def load_data():
    db = get_db()
    if db is not None:
        try:
            funcionarios = list(db.funcionarios.find({}, {'_id': 0}))
            vagas = list(db.vagas.find({}, {'_id': 0}))
            return funcionarios, vagas
        except Exception as e:
            current_app.config['DB_ERROR'] = str(e)
            return [], []
    return [], []

@main.route('/')
@main.route('/dashboard')
def dashboard():
    funcionarios, vagas = load_data()
    db_error = current_app.config.get('DB_ERROR')
    query = request.args.get('busca_habilidade', '')
    filtro_area = request.args.get('area', '')
    filtro_cargo = request.args.get('cargo', '')
    ordem = request.args.get('ordem', '')
    
    if query:
        funcionarios_filtrados = []
        for f in funcionarios:
            habilidades_declaradas = f.get('habilidades_declaradas', [])
            habilidades_descobertas = [h['skill'] for h in f.get('habilidades_descobertas', [])]
            habilidades = habilidades_declaradas + habilidades_descobertas
            if any(query.lower() in habilidade.lower() for habilidade in habilidades):
                funcionarios_filtrados.append(f)
        funcionarios = funcionarios_filtrados
    if filtro_cargo:
        funcionarios = [f for f in funcionarios if filtro_cargo.lower() in f.get('cargo', '').lower()]
    def normalizar_skill(s):
        s = s or ''
        s = re.sub(r"\s*\(.*?\)\s*", '', s)
        return s.strip().lower()

    total_colaboradores = len(funcionarios)
    contagem = {}
    display_map = {}
    for f in funcionarios:
        habilidades_declaradas = f.get('habilidades_declaradas', [])
        habilidades_descobertas = [h['skill'] for h in f.get('habilidades_descobertas', [])]
        for s in habilidades_declaradas + habilidades_descobertas:
            k = normalizar_skill(s)
            if k:
                contagem[k] = contagem.get(k, 0) + 1
                if k not in display_map:
                    display_map[k] = s
    top_skills = sorted(({"skill": display_map[k], "count": c} for k, c in contagem.items()), key=lambda x: x["count"], reverse=True)[:5]

    def melhor_compatibilidade(funcionario):
        habilidades_decl = funcionario.get('habilidades_declaradas', [])
        habilidades_desc = [h['skill'] for h in funcionario.get('habilidades_descobertas', [])]
        mapa_func = {normalizar_skill(s) for s in (habilidades_decl + habilidades_desc)}
        melhor = 0
        melhor_area = ''
        for vaga in vagas:
            mapa_vaga = {normalizar_skill(s) for s in vaga.get('habilidades_requeridas', [])}
            if not mapa_vaga:
                continue
            comuns = mapa_func.intersection(mapa_vaga)
            comp = round((len(comuns) / len(mapa_vaga)) * 100)
            if comp > melhor:
                melhor = comp
                melhor_area = vaga.get('area', '')
        return melhor, melhor_area

    anotados = []
    for f in funcionarios:
        comp, area_top = melhor_compatibilidade(f)
        f['melhor_compatibilidade'] = comp
        f['area_top'] = area_top
        anotados.append(f)

    if filtro_area:
        anotados = [f for f in anotados if f.get('area_top', '').lower() == filtro_area.lower()]

    if ordem == 'compat':
        anotados.sort(key=lambda x: x.get('melhor_compatibilidade', 0), reverse=True)
    else:
        anotados.sort(key=lambda x: x.get('nome', ''))

    def extrair_skills_texto(texto):
        encontrados = re.findall(r"\b(" + "|".join(re.escape(s) for s in SKILLS_CONHECIDAS) + r")\b", texto.lower())
        return sorted({normalizar_skill(s) for s in encontrados})

    def skill_gap_projetos():
        try:
            with open('app/data/projetos.json', 'r', encoding='utf-8') as f:
                projetos = json.load(f)
        except FileNotFoundError:
            return []
        mapa_func = {f['id']: f for f in funcionarios}
        resultado = []
        for proj in projetos:
            skills_req = set()
            for t in proj.get('tarefas', []):
                skills_req.update(extrair_skills_texto(t.get('descricao', '')))
            participantes = proj.get('participantes', [])
            cobertura = []
            for s in sorted(skills_req):
                total = len(participantes) or 1
                tem = 0
                for pid in participantes:
                    f = mapa_func.get(pid)
                    if not f:
                        continue
                    decl = {normalizar_skill(x) for x in f.get('habilidades_declaradas', [])}
                    desc = {normalizar_skill(x.get('skill')) for x in f.get('habilidades_descobertas', [])}
                    if s in decl or s in desc:
                        tem += 1
                cobertura.append({"skill": s.capitalize(), "percent": round((tem / total) * 100)})
            resultado.append({
                "id_projeto": proj.get('id_projeto'),
                "nome": proj.get('nome_projeto'),
                "cobertura": cobertura
            })
        return resultado

    gap = skill_gap_projetos()

    return render_template('dashboard.html', funcionarios=anotados, query=query, total_colaboradores=total_colaboradores, top_skills=top_skills, filtro_area=filtro_area, filtro_cargo=filtro_cargo, ordem=ordem, gap=gap, db_error=db_error)

@main.route('/perfil/<int:id>')
def perfil(id):
    funcionarios, vagas = load_data()
    funcionario = next((f for f in funcionarios if f['id'] == id), None)
    
    if funcionario is None:
        abort(404)

    def normalizar_skill(s):
        s = s or ''
        s = re.sub(r"\s*\(.*?\)\s*", '', s)  # remove níveis/parenteses
        return s.strip().lower()

    habilidades_decl = funcionario.get('habilidades_declaradas', [])
    habilidades_desc = [h['skill'] for h in funcionario.get('habilidades_descobertas', [])]

    mapa_func = {}
    for s in habilidades_decl + habilidades_desc:
        mapa_func[normalizar_skill(s)] = s

    recomendacoes = []
    for vaga in vagas:
        habilidades_req = vaga.get('habilidades_requeridas', [])
        mapa_vaga = {normalizar_skill(s): s for s in habilidades_req}
        chaves_func = set(mapa_func.keys())
        chaves_vaga = set(mapa_vaga.keys())
        chaves_comuns = chaves_func.intersection(chaves_vaga)
        
        if chaves_comuns:
            compatibilidade = round((len(chaves_comuns) / len(chaves_vaga)) * 100) if chaves_vaga else 0
            limiar = request.args.get('limiar', default=COMPATIBILIDADE_LIMIAR, type=int)
            if compatibilidade > limiar:
                recomendacoes.append({
                    "id": vaga['id'],
                    "titulo": vaga['titulo'],
                    "compatibilidade": compatibilidade,
                    "habilidades_em_comum": [mapa_vaga[k] for k in sorted(chaves_comuns)],
                    "habilidades_a_desenvolver": [mapa_vaga[k] for k in sorted(chaves_vaga - chaves_func)]
                })
    
    recomendacoes.sort(key=lambda x: x['compatibilidade'], reverse=True)
    
    funcionario['recomendacoes_vagas'] = recomendacoes

    return render_template('perfil.html', funcionario=funcionario)

@main.route('/atualizar_skills')
def atualizar_skills():
    db = get_db()
    if db is None:
        return redirect(url_for('main.dashboard', erro_conexao=1))
    try:
        projetos = list(db.projetos.find({}, {'_id': 0}))
        skills = set(SKILLS_CONHECIDAS)
        regex = re.compile(r"\b(" + "|".join(re.escape(s) for s in skills) + r")\b", re.IGNORECASE)
        from datetime import datetime
        hoje = datetime.utcnow().strftime('%Y-%m-%d')
        for projeto in projetos:
            participantes = projeto.get('participantes', [])
            encontrados = set()
            for tarefa in projeto.get('tarefas', []):
                texto = (tarefa.get('descricao', '') or '').lower()
                for m in regex.findall(texto):
                    encontrados.add(m.strip())
            for pid in participantes:
                doc = db.funcionarios.find_one({'id': pid})
                if not doc:
                    continue
                existentes = doc.get('habilidades_descobertas', []) or []
                existentes_map = { (e.get('skill') or '').strip().lower() for e in existentes if isinstance(e, dict) }
                novos = []
                for s in sorted(encontrados):
                    k = re.sub(r"\s*\(.*?\)\s*", '', s).strip().lower()
                    if k and k not in existentes_map:
                        novos.append({'skill': k.capitalize(), 'origem': 'NLP (projetos)', 'data': hoje})
                if novos:
                    db.funcionarios.update_one({'id': pid}, {'$set': {'habilidades_descobertas': existentes + novos}})
        return redirect(url_for('main.dashboard', updated=1))
    except Exception:
        return redirect(url_for('main.dashboard', updated=0))

@main.route('/graficos')
def graficos():
    funcionarios, vagas = load_data()
    db_error = None
    db = get_db()
    if db is None:
        db_error = 'Erro de conexão com MongoDB.'
    def normalizar_skill(s):
        s = s or ''
        s = re.sub(r"\s*\(.*?\)\s*", '', s)
        return s.strip().lower()

    contagem = {}
    display_map = {}
    for f in funcionarios:
        habilidades_declaradas = f.get('habilidades_declaradas', [])
        habilidades_descobertas = [h['skill'] for h in f.get('habilidades_descobertas', [])]
        for s in habilidades_declaradas + habilidades_descobertas:
            k = normalizar_skill(s)
            if k:
                contagem[k] = contagem.get(k, 0) + 1
                if k not in display_map:
                    display_map[k] = s
    top_labels = [display_map[k] for k, _ in sorted(contagem.items(), key=lambda x: x[1], reverse=True)[:10]]
    top_values = [contagem[normalizar_skill(lbl)] for lbl in top_labels]

    area_counts = {}
    compat_values = []
    for f in funcionarios:
        habilidades_decl = f.get('habilidades_declaradas', [])
        habilidades_desc = [h['skill'] for h in f.get('habilidades_descobertas', [])]
        mapa_func = {normalizar_skill(s) for s in (habilidades_decl + habilidades_desc)}
        melhor = 0
        melhor_area = ''
        for vaga in vagas:
            mapa_vaga = {normalizar_skill(s) for s in vaga.get('habilidades_requeridas', [])}
            if not mapa_vaga:
                continue
            comuns = mapa_func.intersection(mapa_vaga)
            comp = round((len(comuns) / len(mapa_vaga)) * 100)
            if comp > melhor:
                melhor = comp
                melhor_area = vaga.get('area', '')
        compat_values.append(melhor)
        if melhor_area:
            area_counts[melhor_area] = area_counts.get(melhor_area, 0) + 1

    bins = [0, 20, 40, 60, 80, 100]
    dist_labels = ["0-20", "21-40", "41-60", "61-80", "81-100"]
    dist_values = [0, 0, 0, 0, 0]
    for v in compat_values:
        if v <= 20:
            dist_values[0] += 1
        elif v <= 40:
            dist_values[1] += 1
        elif v <= 60:
            dist_values[2] += 1
        elif v <= 80:
            dist_values[3] += 1
        else:
            dist_values[4] += 1

    projetos = list(db.projetos.find({}, {'_id': 0})) if db else []
    mapa_func = {f['id']: f for f in funcionarios}
    proj_labels = []
    proj_values = []
    for proj in projetos:
        participantes = proj.get('participantes', [])
        total = len(participantes) or 1
        skills_req = set()
        for t in proj.get('tarefas', []):
            encontrados = re.findall(r"\b(" + "|".join(re.escape(s) for s in SKILLS_CONHECIDAS) + r")\b", t.get('descricao', '').lower())
            for s in encontrados:
                skills_req.add(normalizar_skill(s))
        if not skills_req:
            continue
        cobertura_percent = []
        for s in skills_req:
            tem = 0
            for pid in participantes:
                f = mapa_func.get(pid)
                if not f:
                    continue
                decl = {normalizar_skill(x) for x in f.get('habilidades_declaradas', [])}
                desc = {normalizar_skill(x.get('skill')) for x in f.get('habilidades_descobertas', [])}
                if s in decl or s in desc:
                    tem += 1
            cobertura_percent.append((tem / total) * 100)
        media = round(sum(cobertura_percent) / len(cobertura_percent))
        proj_labels.append(proj.get('nome_projeto'))
        proj_values.append(media)

    return render_template(
        'graficos.html',
        top_labels=top_labels,
        top_values=top_values,
        area_labels=list(area_counts.keys()),
        area_values=list(area_counts.values()),
        dist_labels=dist_labels,
        dist_values=dist_values,
        proj_labels=proj_labels,
        proj_values=proj_values,
        db_error=db_error
    )

@main.route('/styleguide')
def styleguide():
    return render_template('styleguide.html')

@main.route('/novo_usuario', methods=['GET', 'POST'])
def novo_usuario():
    if request.method == 'POST':
        nome = request.form.get('nome', '').strip()
        cargo = request.form.get('cargo', '').strip()
        email = request.form.get('email', '').strip()
        habilidades = request.form.get('habilidades_declaradas', '').strip()
        habilidades_lista = [h.strip() for h in habilidades.split(',') if h.strip()] if habilidades else []

        db = get_db()
        if db is None:
            return render_template('upload_usuarios.html', erro='Sem conexão com MongoDB. Configure MONGODB_URI/MONGODB_DB e reinicie.'), 500

        max_doc = db.funcionarios.find_one(sort=[('id', -1)])
        novo_id = (max_doc['id'] + 1) if max_doc and 'id' in max_doc else 1
        db.funcionarios.insert_one({
            'id': novo_id,
            'nome': nome,
            'cargo': cargo,
            'email': email,
            'habilidades_declaradas': habilidades_lista,
            'habilidades_descobertas': []
        })
        return redirect(url_for('main.dashboard', created=1))
    return render_template('novo_usuario.html')

@main.route('/upload_usuarios', methods=['GET', 'POST'])
def upload_usuarios():
    if request.method == 'POST':
        file = request.files.get('arquivo')
        if not file or not file.filename.lower().endswith('.xlsx'):
            return redirect(url_for('main.upload_usuarios', erro=1))
        wb = load_workbook(filename=file, read_only=True)
        ws = wb['Funcionarios'] if 'Funcionarios' in wb.sheetnames else wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return redirect(url_for('main.upload_usuarios', erro=1))
        header = [str(h).strip().lower() if h is not None else '' for h in rows[0]]
        idx = {k: header.index(k) if k in header else -1 for k in ['id','nome','cargo','email','habilidades_declaradas']}
        novos = []
        for r in rows[1:]:
            nome = (r[idx['nome']] if idx['nome']!=-1 else '') or ''
            cargo = (r[idx['cargo']] if idx['cargo']!=-1 else '') or ''
            email = (r[idx['email']] if idx['email']!=-1 else '') or ''
            habilidades_raw = (r[idx['habilidades_declaradas']] if idx['habilidades_declaradas']!=-1 else '') or ''
            habilidades = [h.strip() for h in str(habilidades_raw).split(',') if str(h).strip()] if habilidades_raw else []
            if nome and cargo and email:
                novos.append({'nome': nome, 'cargo': cargo, 'email': email, 'habilidades_declaradas': habilidades})
        db = get_db()
        if db is None:
            return render_template('upload_usuarios.html', erro='Sem conexão com MongoDB. Configure MONGODB_URI/MONGODB_DB e reinicie.'), 500
        existentes = set(doc.get('email') for doc in db.funcionarios.find({}, {'email': 1, '_id': 0}))
        max_doc = db.funcionarios.find_one(sort=[('id', -1)])
        next_id = (max_doc['id'] + 1) if max_doc and 'id' in max_doc else 1
        docs = []
        importados = 0
        for n in novos:
            if n['email'] in existentes:
                continue
            docs.append({
                'id': next_id,
                'nome': n['nome'],
                'cargo': n['cargo'],
                'email': n['email'],
                'habilidades_declaradas': n['habilidades_declaradas'],
                'habilidades_descobertas': []
            })
            next_id += 1
            importados += 1
        if docs:
            db.funcionarios.insert_many(docs)
        return redirect(url_for('main.dashboard', importados=importados))
    return render_template('upload_usuarios.html')

@main.route('/plano_carreira/<int:id>', methods=['POST'])
def plano_carreira(id):
    funcionarios, vagas = load_data()
    funcionario = next((f for f in funcionarios if f['id'] == id), None)
    if not funcionario:
        abort(404)

    def normalizar_skill(s):
        s = s or ''
        s = re.sub(r"\s*\(.*?\)\s*", '', s)
        return s.strip().lower()

    habilidades_decl = funcionario.get('habilidades_declaradas', [])
    habilidades_desc = [h['skill'] for h in funcionario.get('habilidades_descobertas', [])]
    mapa_func = {}
    for s in habilidades_decl + habilidades_desc:
        mapa_func[normalizar_skill(s)] = s

    recomendacoes = []
    for vaga in vagas:
        habilidades_req = vaga.get('habilidades_requeridas', [])
        mapa_vaga = {normalizar_skill(s): s for s in habilidades_req}
        chaves_func = set(mapa_func.keys())
        chaves_vaga = set(mapa_vaga.keys())
        chaves_comuns = chaves_func.intersection(chaves_vaga)
        if chaves_comuns:
            compatibilidade = round((len(chaves_comuns) / len(chaves_vaga)) * 100) if chaves_vaga else 0
            recomendacoes.append({
                "id": vaga['id'],
                "titulo": vaga['titulo'],
                "compatibilidade": compatibilidade,
                "habilidades_em_comum": [mapa_vaga[k] for k in sorted(chaves_comuns)],
                "habilidades_a_desenvolver": [mapa_vaga[k] for k in sorted(chaves_vaga - chaves_func)]
            })
    recomendacoes.sort(key=lambda x: x['compatibilidade'], reverse=True)

    api_key = os.getenv('OPENAI_API_KEY')
    plano = None
    erro_ai = None
    if not api_key:
        erro_ai = 'Chave de API ausente (OPENAI_API_KEY).'
    else:
        try:
            client = OpenAI(api_key=api_key)
            mensagens = [
                {"role": "system", "content": "Você é um assistente de carreira. Gere plano prático e acionável."},
                {"role": "user", "content": (
                    f"Nome: {funcionario.get('nome')}\n"
                    f"Cargo atual: {funcionario.get('cargo')}\n"
                    f"Habilidades declaradas: {', '.join(habilidades_decl) or 'Nenhuma'}\n"
                    f"Habilidades descobertas: {', '.join(habilidades_desc) or 'Nenhuma'}\n"
                    f"Recomendações internas: "
                    + ", ".join([f"{r['titulo']} ({r['compatibilidade']}%)" for r in recomendacoes[:5]])
                    + "\nCrie um plano de carreira de 6-12 meses com: metas mensais, projetos internos sugeridos, habilidades a desenvolver e cursos/trilhas curtas. Linguagem objetiva em bullet points.\n"
                    + "Inclua recomendações de cursos com plataforma (somente: Alura, Data Science Academy, Udemy, Microsoft Learning), carga horária estimada e nível (iniciante/intermediário/avançado). Não sugira plataformas fora dessa lista."
                )}
            ]
            resp = client.chat.completions.create(model=os.getenv('OPENAI_MODEL', 'gpt-5-nano'), messages=mensagens)
            plano = resp.choices[0].message.content
        except Exception as e:
            erro_ai = str(e)

    funcionario['recomendacoes_vagas'] = recomendacoes
    if plano:
        funcionario['plano_carreira'] = plano
    if erro_ai:
        funcionario['plano_carreira_erro'] = erro_ai

    return render_template('perfil.html', funcionario=funcionario)

@main.app_errorhandler(404)
def page_not_found(e):
    return render_template('404.html'), 404
